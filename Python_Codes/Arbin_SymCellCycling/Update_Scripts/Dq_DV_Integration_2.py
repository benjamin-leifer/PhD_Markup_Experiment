#!/usr/bin/env python3
"""
dqdv_by_cell.py

Standalone dQ/dV plotting for Arbin exports.

Features
--------
- Robust loader for Arbin Excel/CSV with common column-name variants
- Cycle and half-cycle selection (charge/discharge)
- Fixed-width voltage binning (default 3 mV)
- dQ/dV via raw differencing or Savitzky–Golay smoothing (if SciPy available)
- Plot overlay (one axis) or per-cell subplots (grouped by cell code prefix)
- Small CLI

Assumptions
-----------
- "Cell code" is the prefix in the filename before the first underscore: e.g.
  "BL-LL-HL01_RT_Channel_37_Wb_1.xlsx" -> "BL-LL-HL01"

Author: you + ChatGPT
"""

from __future__ import annotations
import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Tuple, Optional

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

# ---- Optional: Savitzky–Golay smoothing (if SciPy installed) ----------------
try:
    from scipy.signal import savgol_filter  # type: ignore
    _HAS_SCIPY = True
except Exception:
    _HAS_SCIPY = False

# ============================== Configuration =================================

# Defaults (can be overridden by CLI)
DATA_DIR = Path(r"C:\Users\benja\Downloads\Dilute THF Data\10_14")
FILES: List[Path] = []              # if empty, we'll glob DATA_DIR instead
BIN_W = 0.003                       # 3 mV bins
CYCLE = 1
CHARGE = True                       # True: charge, False: discharge
DQDV_SMOOTH = False                 # use Savitzky–Golay if available

# ============================== I/O Utilities =================================

def list_files(
    data_dir: Path,
    file_glob: Optional[str] = None,
    recursive: bool = True
) -> List[Path]:
    """Find candidate files (xlsx/csv/mpt). Optionally recurse into subfolders."""
    patterns = []
    if file_glob:
        patterns.append(file_glob)

    # Always include common types
    default_exts = ["*.xlsx", "*.xlsm", "*.xls", "*.csv", "*.mpt", "*.txt"]
    patterns.extend(default_exts)

    globber = data_dir.rglob if recursive else data_dir.glob
    files: List[Path] = []
    seen = set()
    for pat in patterns:
        for p in globber(pat):
            rp = p.resolve()
            if rp not in seen and p.is_file():
                files.append(p)
                seen.add(rp)
    return sorted(files)



def _find_voltage_like_column(columns):
    """Return the first column name that looks like voltage."""
    import re
    pats = [
        r"voltage", r"cell\s*volt", r"\bewe\b", r"\becell\b",
        r"\bv\s*\(v\)", r"\bv\s*\[v\]", r"\bv\b", r"e\(v\)"
    ]
    lowmap = {str(c).strip().lower(): c for c in columns}
    for p in pats:
        rx = re.compile(p, re.IGNORECASE)
        for low, orig in lowmap.items():
            if rx.search(low):
                return orig
    # also accept exact single-letter "v" or variants
    for c in columns:
        if str(c).strip().lower() in {"v", "v (v)", "voltage (v)", "voltage[v]"}:
            return c
    return None


def _header_row_values(ws):
    """Return header row values (as strings) for a worksheet, or [] if empty."""
    row_iter = ws.iter_rows(max_row=1, values_only=True)
    try:
        first = next(row_iter)
    except StopIteration:
        return []
    return [("" if v is None else str(v)).strip() for v in first]


def _looks_like_voltage(headers):
    """Heuristic: does the header row contain a voltage-like column?"""
    import re
    pats = [
        r"\bvoltage\b", r"cell\s*volt", r"\bewe\b", r"\becell\b",
        r"^\s*v\s*$", r"\bv\s*\(v\)", r"\bv\s*\[v\]", r"e\(v\)"
    ]
    low = [h.lower() for h in headers]
    for p in pats:
        rx = re.compile(p, re.IGNORECASE)
        if any(rx.search(h) for h in low):
            return True
    # extra loose checks
    for h in low:
        if h.replace(" ", "") in {"voltage(v)", "cellvoltage(v)", "voltage"}:
            return True
    return False


def read_any(fp: Path) -> pd.DataFrame:
    """
    Read Arbin-like export (xlsx/csv/mpt). For Excel, pick the best sheet by
    scanning ONLY header rows (fast), then parse that single sheet with pandas.
    """
    suffix = fp.suffix.lower()
    if suffix in (".xlsx", ".xls", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        sheetnames = wb.sheetnames

        # 1) Prefer sheets with common channel/data names
        name_priority = []
        for sn in sheetnames:
            s = sn.lower()
            # Higher priority first
            score = 0
            if "channel" in s and "data" in s: score += 3
            if "data" in s: score += 2
            if "sheet" in s: score -= 1
            name_priority.append((score, sn))
        # Try in priority order, but only reading header rows
        chosen = None
        fallback_first_voltage = None
        fallback_any = None

        for _, sn in sorted(name_priority, key=lambda x: -x[0]):
            ws = wb[sn]
            headers = _header_row_values(ws)
            if headers:
                if _looks_like_voltage(headers):
                    chosen = sn
                    break
                if fallback_first_voltage is None and len(headers) >= 3:
                    # mark a reasonable candidate in case none match voltage
                    fallback_first_voltage = sn
            if fallback_any is None:
                fallback_any = sn

        wb.close()

        sheet_to_parse = chosen or fallback_first_voltage or fallback_any or sheetnames[0]
        # 2) Now parse ONLY the chosen sheet fully
        return pd.read_excel(fp, sheet_name=sheet_to_parse, engine="openpyxl")

    elif suffix in (".csv", ".txt", ".mpt"):
        return pd.read_csv(fp)
    else:
        raise ValueError(f"Unsupported file type: {fp}")





# ============================== Core Processing ===============================

@dataclass
class ColumnMap:
    voltage: str
    current: Optional[str]
    cycle: Optional[str]
    charge_cap: Optional[str]
    discharge_cap: Optional[str]
    step: Optional[str]


def detect_columns(df: pd.DataFrame) -> ColumnMap:
    """Detect likely columns from common Arbin variants, with regex fallbacks."""
    import re
    cols = list(df.columns)

    def pick_one(name_list: Iterable[str], regex_list: Iterable[str] = ()):
        lowmap = {str(c).strip().lower(): c for c in cols}
        for name in name_list:
            key = name.lower()
            if key in lowmap:
                return lowmap[key]
        for k, orig in lowmap.items():
            for name in name_list:
                if name.lower() in k:
                    return orig
        for rx in regex_list:
            r = re.compile(rx, re.IGNORECASE)
            for c in cols:
                if r.search(str(c)):
                    return c
        return None

    voltage = pick_one(
        ["voltage", "voltage(v)", "v", "potential", "ewe(v)", "cell voltage", "cell voltage (v)"],
        [r"\bvoltage\b", r"cell\s*volt", r"\bewe\b", r"\becell\b", r"^\s*v\s*$", r"\bv\s*\(v\)"]
    )
    current = pick_one(
        ["current", "current(a)", "i", "current(mA)", "currenta", "battery current", "current (a)"],
        [r"\bcurrent\b", r"\bi\b", r"amps?\b"]
    )
    cycle   = pick_one(
        ["cycle", "cycle_index", "cycle number", "cycleindex"],
        [r"cycle"]
    )
    ch_cap  = pick_one(
        ["charge_capacity", "charge capacity", "charge_capacity(ah)", "charge_capacity(mAh)", "q_charge",
         "chargecap(ah)", "chg. capacity (mah)", "charge capacity (mah)"],
        [r"charge.*cap", r"\bq[_\s-]*charge\b"]
    )
    dch_cap = pick_one(
        ["discharge_capacity", "discharge capacity", "discharge_capacity(ah)", "discharge_capacity(mAh)", "q_discharge",
         "dischargecap(ah)", "dchg. capacity (mah)", "discharge capacity (mah)"],
        [r"discharge.*cap", r"\bq[_\s-]*discharge\b"]
    )
    step    = pick_one(
        ["step", "step_index", "step time index", "stepnumber"],
        [r"\bstep(\s*time)?(\s*index)?\b"]
    )

    if voltage is None:
        print("[detect_columns] Could not find a voltage column. Headers were:")
        for c in df.columns:
            print("  -", repr(c))
        raise KeyError("Could not detect a voltage column (e.g., 'Voltage(V)').")

    return ColumnMap(
        voltage=voltage,
        current=current,
        cycle=cycle,
        charge_cap=ch_cap,
        discharge_cap=dch_cap,
        step=step
    )


def select_half_cycle(df: pd.DataFrame, cols: ColumnMap, charge: bool) -> pd.DataFrame:
    """
    Select rows corresponding to charge or discharge half-cycle using
    (1) explicit capacity columns if present or
    (2) current sign as a fallback.
    """
    work = df.copy()

    if cols.cycle and cols.cycle in work:
        # Ensure cycle is integer-like
        work[cols.cycle] = pd.to_numeric(work[cols.cycle], errors="coerce").astype("Int64")

    # Prefer explicit capacity columns to infer half-cycle segments
    if cols.charge_cap and cols.charge_cap in work and cols.discharge_cap and cols.discharge_cap in work:
        ch = pd.to_numeric(work[cols.charge_cap], errors="coerce")
        dch = pd.to_numeric(work[cols.discharge_cap], errors="coerce")
        if charge:
            mask = (ch.notna()) & (ch.diff().fillna(0) >= 0)
        else:
            mask = (dch.notna()) & (dch.diff().fillna(0) >= 0)
        work = work[mask]
    elif cols.current and cols.current in work:
        # Fall back to current sign: charge -> I > 0 (convention varies!)
        i = pd.to_numeric(work[cols.current], errors="coerce")
        mask = i > 0 if charge else i < 0
        # If nearly all False (convention reversed), flip rule
        if mask.sum() < len(mask) * 0.1:
            mask = ~mask
        work = work[mask].copy()
    else:
        # If we cannot decide, just return all rows (will be filtered by cycle later)
        work = work.copy()

    return work


def load_arbin(fp: Path, cycle: int, charge: bool) -> pd.DataFrame:
    """
    Load an Arbin file and return a dataframe with columns:
    - 'V'   : Voltage (V)
    - 'QmAh': Capacity in Ah (YES, Ah) to match older scripts that later *x1000*
    """
    raw = read_any(fp)
    cols = detect_columns(raw)
    df = select_half_cycle(raw, cols, charge=charge)

    # Filter by cycle if available
    if cols.cycle and cols.cycle in df:
        df = df[df[cols.cycle] == cycle]

    # Build a capacity trace:
    # If charge/discharge capacity column exists, use it; else integrate I dt if available; else monotonic index proxy disabled.
    q_ah = None
    if cols.charge_cap and cols.charge_cap in df and cols.discharge_cap and cols.discharge_cap in df:
        # Choose correct column by half-cycle, convert to Ah if needed
        if charge:
            q = pd.to_numeric(df[cols.charge_cap], errors="coerce")
        else:
            q = pd.to_numeric(df[cols.discharge_cap], errors="coerce")
        # Convert to Ah if column looks like mAh
        # Heuristic: if max > 20, assume mAh and scale to Ah
        if q.max(skipna=True) is not None and float(q.max(skipna=True)) > 20.0:
            q_ah = q / 1000.0
        else:
            q_ah = q
    else:
        # Fallback: if we have current, integrate roughly by index (no dt in export reliably)
        # Many Arbin exports also include "Test Time (s)" but columns vary widely.
        # We'll simply cumulative-sum current in A over arbitrary dt=1s if 'Test Time(s)' found; else return NaN
        time_col = None
        for cand in ["Test Time(s)", "test time(s)", "Time(s)", "time(s)"]:
            if cand in df.columns:
                time_col = cand
                break
        if cols.current and cols.current in df and time_col:
            i = pd.to_numeric(df[cols.current], errors="coerce").fillna(0.0)  # A
            t = pd.to_numeric(df[time_col], errors="coerce").fillna(method="ffill").fillna(0.0)  # s
            dt = np.r_[0.0, np.diff(t.values)]
            q_as = np.cumsum(i.values * dt)  # A*s
            q_ah = pd.Series(q_as / 3600.0, index=df.index)
        else:
            # Last resort: use normalized cumulative index (unitless); still allows dQ/dV shape
            q_ah = pd.Series(np.linspace(0, 1, len(df)), index=df.index)

    v = pd.to_numeric(df[cols.voltage], errors="coerce")

    out = pd.DataFrame({
        "V": v.values,
        "QmAh": q_ah.values  # keep in Ah; caller may multiply by 1000 to get mAh
    }).dropna()

    # Remove non-finite and ensure monotonic by V for binning
    out = out[np.isfinite(out["V"].values) & np.isfinite(out["QmAh"].values)]
    out = out.sort_values("V").reset_index(drop=True)
    return out


def fixed_bin(df: pd.DataFrame, bin_w: float) -> pd.DataFrame:
    """
    Bin by fixed voltage width.
    Returns DataFrame with columns ['Vmid', 'QmAh_mean'] where QmAh_mean is the mean within each bin.
    """
    v = df["V"].values
    q = df["QmAh"].values
    if len(v) < 3:
        return pd.DataFrame(columns=["Vmid", "QmAh_mean"])

    vmin, vmax = np.nanmin(v), np.nanmax(v)
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmax <= vmin:
        return pd.DataFrame(columns=["Vmid", "QmAh_mean"])

    edges = np.arange(vmin, vmax + bin_w, bin_w)
    if len(edges) < 3:
        return pd.DataFrame(columns=["Vmid", "QmAh_mean"])

    idx = np.digitize(v, edges) - 1
    # aggregate by bin index
    data = {}
    for b in range(idx.min(), idx.max() + 1):
        mask = idx == b
        if mask.sum() >= 1:
            vbin = v[mask]
            qbin = q[mask]
            data[b] = (np.nanmean(vbin), np.nanmean(qbin))
    if not data:
        return pd.DataFrame(columns=["Vmid", "QmAh_mean"])

    arr = np.array(list(data.values()))
    return pd.DataFrame({"Vmid": arr[:, 0], "QmAh_mean": arr[:, 1]}).sort_values("Vmid").reset_index(drop=True)


def raw_dqdv(df_bin: pd.DataFrame) -> Tuple[np.ndarray, np.ndarray]:
    """
    Compute raw dQ/dV from binned data.
    Input df_bin: columns ['Vmid', 'QmAh_mean'] with Q in Ah (will return dQ/dV in mAh/V if multiplied upstream).
    """
    v = df_bin["Vmid"].values
    q = df_bin["QmAh_mean"].values
    if len(v) < 3:
        return v, np.zeros_like(v)

    dq = np.diff(q)  # Ah
    dv = np.diff(v)  # V
    with np.errstate(divide="ignore", invalid="ignore"):
        dqdv = dq / dv
    # center between edges
    v_mid = 0.5 * (v[1:] + v[:-1])
    # convert Ah/V -> mAh/V
    return v_mid, dqdv * 1000.0


def savgol_dqdv(df_bin: pd.DataFrame, window: int = 11, poly: int = 3) -> Tuple[np.ndarray, np.ndarray]:
    """
    Compute dQ/dV with Savitzky–Golay differentiation (requires SciPy).
    """
    v = df_bin["Vmid"].values
    q = df_bin["QmAh_mean"].values
    n = len(v)
    if n < max(5, poly + 2):
        return raw_dqdv(df_bin)

    if not _HAS_SCIPY:
        print("[warn] SciPy not installed; falling back to raw differencing.")
        return raw_dqdv(df_bin)

    # ensure window odd and <= n
    if window % 2 == 0:
        window += 1
    window = max(5, min(window, n - (1 - n % 2)))

    # Savitzky–Golay derivative wrt sample index, then divide by dv/dx
    q_sg = savgol_filter(q, window_length=window, polyorder=poly, deriv=1, delta=1.0, mode="interp")
    # approximate dv/dx
    dv_dx = np.gradient(v)
    with np.errstate(divide="ignore", invalid="ignore"):
        dqdv = q_sg / dv_dx
    return v, dqdv * 1000.0  # mAh/V


# ============================== Plotting Helpers ==============================

def _iter_files_from_globals(
    data_dir: Optional[Path] = None,
    file_glob: Optional[str] = None,
    recursive: bool = True
) -> List[Path]:
    dd = Path(data_dir or DATA_DIR)
    if FILES:
        return [p if isinstance(p, Path) else dd / str(p) for p in FILES]
    return list_files(dd, file_glob=file_glob, recursive=recursive)



def _plot_group_on_ax(ax: plt.Axes, cell_code: str, files: List[Path],
                      cycle: int, charge: bool, bin_w: float, smooth: bool) -> None:
    """
    Plot all replicates for a given cell code onto one axis.
    Includes progress updates every 10 files.
    """
    drew_any = False
    total = len(files)
    for i, f in enumerate(files, start=1):
        # ---- Progress print line ----
        if i % 10 == 1 or i == total:
            print(f"[{cell_code}] Processing file {i}/{total}: {f.name}", flush=True)
        # -----------------------------

        try:
            df_raw = load_arbin(f, cycle, charge)
            df_bin = fixed_bin(df_raw, bin_w)
            if len(df_bin) < 5:
                continue
            v_mid, y = (savgol_dqdv(df_bin) if smooth else raw_dqdv(df_bin))
            ax.plot(v_mid, y, lw=1.6, label=f.stem)
            drew_any = True
        except Exception as e:
            print(f"[{cell_code}] Skipped {f.name}: {e}")

    if drew_any:
        ax.set_ylabel("dQ/dV (mAh V$^{-1}$)")
        ax.set_title(f"{cell_code} – Cycle {cycle}, {'Charge' if charge else 'Discharge'}")
        ax.legend(fontsize="xx-small", frameon=True)



def plot_dqdv_cells(
    per_cell: bool = False,
    savepath: Optional[Path] = None,
    data_dir: Optional[Path] = None,
    file_glob: Optional[str] = None,
    cycle: Optional[int] = None,
    charge: Optional[bool] = None,
    bin_w: Optional[float] = None,
    smooth: Optional[bool] = None,
    recursive: bool = True,
) -> None:
    """
    Plot dQ/dV curves grouped by cell code (prefix before first underscore).

    Modes:
        per_cell=False  -> overlay all replicates for every cell code on one axis
        per_cell=True   -> each cell code gets its own subplot

    Uses global defaults unless overridden.
    """

    import itertools
    from matplotlib import cm
    from matplotlib.lines import Line2D

    cyc = CYCLE if cycle is None else cycle
    chg = CHARGE if charge is None else charge
    bw  = BIN_W if bin_w is None else bin_w
    sm  = DQDV_SMOOTH if smooth is None else smooth

    files = _iter_files_from_globals(data_dir=data_dir, file_glob=file_glob, recursive=recursive)
    if not files:
        print(f"[dQ/dV] No files found.\n  dir: {Path(data_dir or DATA_DIR).resolve()}\n"
              f"  glob: {file_glob or '(default patterns: *.xlsx, *.xlsm, *.xls, *.csv, *.mpt, *.txt)'}\n"
              f"  recursive: {recursive}")
        return

    # Group by cell code (prefix before first underscore)
    grouped: dict[str, List[Path]] = {}
    for f in files:
        code = f.stem.split("_")[0]
        grouped.setdefault(code, []).append(f)

    # --- PER-CELL MODE -------------------------------------------------------
    if per_cell:
        print(f"[dQ/dV] Starting plot for {len(grouped)} cell codes "
              f"({sum(len(v) for v in grouped.values())} total files)", flush=True)

        n = len(grouped)
        fig, axes = plt.subplots(n, 1, figsize=(7.8, 3.0 * max(1, n)), sharex=True)
        if n == 1:
            axes = [axes]

        for ax, (cc, flist) in zip(axes, grouped.items()):
            _plot_group_on_ax(ax, cc, flist, cyc, chg, bw, sm)
        axes[-1].set_xlabel("Voltage (V)")
        plt.tight_layout()
        if savepath:
            plt.savefig(savepath, dpi=300, bbox_inches="tight")
        plt.show()
        return

    # --- OVERLAY MODE (all replicates per cell code) -------------------------
    fig, ax = plt.subplots(figsize=(7.8, 4.6))

    codes = list(grouped.keys())
    cmap = cm.get_cmap("tab20", len(codes)) if len(codes) <= 20 else cm.get_cmap("hsv", len(codes))
    code_to_color = {cc: cmap(i) for i, cc in enumerate(codes)}

    plotted_any = False
    for cc, flist in grouped.items():
        color = code_to_color[cc]
        for f in flist:
            try:
                df_raw = load_arbin(f, cyc, chg)
                df_bin = fixed_bin(df_raw, bw)
                if len(df_bin) < 5:
                    continue
                v_mid, y = (savgol_dqdv(df_bin) if sm else raw_dqdv(df_bin))
                # Label each replicate individually
                ax.plot(v_mid, y, lw=1.4, label=f"{cc}: {f.stem}", color=color, alpha=0.9)
                plotted_any = True
            except Exception as e:
                print(f"[overlay] {cc} | {f.name}: {e}")

    ax.set_xlabel("Voltage (V)")
    ax.set_ylabel("dQ/dV (mAh V$^{-1}$)")
    ax.set_title(f"dQ/dV – Cycle {cyc}, {'Charge' if chg else 'Discharge'}")

    if plotted_any:
        # Support for global args to compress legend
        use_rep_legend = globals().get("args", None) and getattr(args, "rep_legend", False)
        if use_rep_legend:
            # Compress legend to one entry per cell code
            handles = [Line2D([0], [0], color=code_to_color[cc], lw=2, label=cc) for cc in codes]
            ax.legend(handles=handles, fontsize="x-small", frameon=True, ncol=2)
        else:
            # Default: list every replicate
            ax.legend(fontsize="xx-small", frameon=True, ncol=1)

    plt.tight_layout()
    if savepath:
        plt.savefig(savepath, dpi=300, bbox_inches="tight")
    plt.show()



# ============================== Legacy-style main =============================

def main() -> None:
    """
    Legacy-like entrypoint: if FILES is set, plot those; else glob DATA_DIR.
    Produces an overlay figure across all found cell codes.
    """
    plot_dqdv_cells(per_cell=False)


# ================================== CLI ======================================

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Plot dQ/dV from Arbin files grouped by cell code.")
    p.add_argument("--dir", type=str, default=None, help="Directory to search (defaults to DATA_DIR).")
    p.add_argument("--glob", type=str, default=None, help="Glob to select files (e.g., 'BL-LL-*.xlsx').")
    p.add_argument("--no-recursive", dest="recursive", action="store_false",
                   help="Do NOT search subfolders (default is recursive).")
    p.set_defaults(recursive=True)
    p.add_argument("--cycle", type=int, default=CYCLE, help=f"Cycle index (default {CYCLE}).")
    p.add_argument("--discharge", action="store_true", help="Plot discharge instead of charge.")
    p.add_argument("--binw", type=float, default=BIN_W, help=f"Voltage bin width in V (default {BIN_W}).")
    p.add_argument("--smooth", action="store_true", help="Use Savitzky–Golay smoothing (requires SciPy).")
    p.add_argument("--overlay", action="store_true", help="Overlay one curve per cell code on a single axis.")
    p.add_argument("--by-cell", default=True, action="store_true", help="One subplot per cell code.")
    p.add_argument("--save", type=str, default=None, help="Path to save figure (PNG).")
    return p



if __name__ == "__main__":
    parser = build_parser()
    args = parser.parse_args()

    # Update globals from CLI
    DATA_DIR = Path(args.dir) if args.dir else DATA_DIR
    CYCLE = int(args.cycle)
    CHARGE = not args.discharge
    BIN_W = float(args.binw)
    DQDV_SMOOTH = bool(args.smooth)

    # Decide plotting mode
    if args.overlay or args.by_cell:
        plot_dqdv_cells(
            per_cell=args.by_cell,
            savepath=Path(args.save) if args.save else None,
            data_dir=DATA_DIR,
            file_glob=args.glob,
            cycle=CYCLE,
            charge=CHARGE,
            bin_w=BIN_W,
            smooth=DQDV_SMOOTH,
            recursive=args.recursive,
        )

    else:
        main()
