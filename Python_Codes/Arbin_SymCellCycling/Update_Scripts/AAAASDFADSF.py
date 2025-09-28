#!/usr/bin/env python3
"""
Streaming Capacity & CE plotter for Arbin Excel files.

- Scans a directory for *.xlsx / *.xls
- Streams each file's channel sheet with openpyxl (fast & memory-light)
- Aggregates per-cycle discharge capacity (mAh), charge capacity (mAh), and CE (%)
- Produces two figures:
  (1) all_cells_capacity_ce.png  — all cells overlaid
  (2) split_by_alpha_capacity_ce.png — subplots split by cell ID prefix (e.g., GY, GZ)
- Also writes metrics_capacity_ce.csv with: Cell, Alpha, Cycle, Qdis_mAh, Qchg_mAh, CE_pct
"""

import re
import math
from pathlib import Path
from typing import List, Tuple, Iterable, Dict
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import csv
from collections import defaultdict


# -----------------------------
# USER CONFIGURATION
# -----------------------------
# Set this to your directory containing the Excel files:
ROOT_DIR = Path(r"C:\Users\benja\Downloads\Temp\C_10 Cycling\2025\08\C_10 Sept Update\Brookhaven Cells")

OUTDIR = ROOT_DIR  # outputs will be written here


# -----------------------------
# Helpers
# -----------------------------
def norm_header(s: str) -> str:
    return s.strip().lower().replace(" ", "").replace("_", "")


def match_columns(header_row: List[str]) -> Dict[str, int]:
    idx_map = {norm_header(h): i for i, h in enumerate(header_row)}
    wants = {
        "cycleindex": None,
        "current(a)": None,
        "chargecapacity(ah)": None,
        "dischargecapacity(ah)": None,
    }
    for key in list(wants.keys()):
        for k, i in idx_map.items():
            if k.startswith(key):
                wants[key] = i
                break
    if any(v is None for v in wants.values()):
        raise KeyError("Required columns not found in sheet header.")
    return wants


def parse_cell_alpha(path: Path) -> Tuple[str, str]:
    stem = path.stem  # e.g., BL-LL-GY01_RT_C3_Cycling_Channel_5_Wb_1
    first_token = stem.split("_")[0]  # BL-LL-GY01
    m = re.search(r"([A-Za-z]{2}\d{2})", first_token)
    cell = m.group(1).upper() if m else first_token[-4:].upper()
    alpha = cell[:2]
    return cell, alpha


def iter_channel_rows(xlsx_path: Path) -> Iterable[List]:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[1]]  # assume 2nd sheet is channel data
    except Exception:
        ws = wb.active
    for row in ws.iter_rows(values_only=True):
        yield list(row)
    wb.close()


def per_file_metrics(xlsx_path: Path) -> List[Tuple[int, float, float, float]]:
    it = iter_channel_rows(xlsx_path)
    header = None
    for row in it:
        if row and any(isinstance(x, str) for x in row):
            header = [str(x) if x is not None else "" for x in row]
            if "Cycle Index" in header or any("Cycle" in h for h in header):
                break
    if header is None:
        raise RuntimeError(f"Could not locate header in {xlsx_path.name}")
    col_idx = match_columns(header)

    current_cycle = None
    chg_vals, chg_vals_pos, dis_vals, dis_vals_neg = [], [], [], []
    results: List[Tuple[int, float, float, float]] = []

    def flush_cycle(cyc):
        if cyc is None:
            return
        chg_max = max(chg_vals_pos) if chg_vals_pos else (max(chg_vals) if chg_vals else 0.0)
        dis_max = max(dis_vals_neg) if dis_vals_neg else (max(dis_vals) if dis_vals else 0.0)
        chg_mAh = 1000.0 * float(chg_max)
        dis_mAh = 1000.0 * float(dis_max)
        ce = 100.0 * (dis_mAh / chg_mAh) if (chg_mAh and chg_mAh != 0 and np.isfinite(chg_mAh)) else 0.0
        results.append((int(cyc), dis_mAh, chg_mAh, ce))

    for row in it:
        try:
            cyc = row[col_idx["cycleindex"]]
            cur = row[col_idx["current(a)"]]
            qchg = row[col_idx["chargecapacity(ah)"]]
            qdis = row[col_idx["dischargecapacity(ah)"]]
        except Exception:
            continue
        if cyc is None:
            continue
        try:
            cyc = int(cyc)
        except Exception:
            continue
        if current_cycle is None:
            current_cycle = cyc
        if cyc != current_cycle:
            flush_cycle(current_cycle)
            current_cycle = cyc
            chg_vals.clear(); chg_vals_pos.clear()
            dis_vals.clear(); dis_vals_neg.clear()
        try:
            qchg = float(qchg) if qchg is not None else np.nan
            qdis = float(qdis) if qdis is not None else np.nan
            cur  = float(cur)  if cur  is not None else np.nan
        except Exception:
            continue
        if np.isfinite(qchg): chg_vals.append(qchg)
        if np.isfinite(qdis): dis_vals.append(qdis)
        if np.isfinite(cur) and np.isfinite(qchg) and cur > 0: chg_vals_pos.append(qchg)
        if np.isfinite(cur) and np.isfinite(qdis) and cur < 0: dis_vals_neg.append(qdis)

    flush_cycle(current_cycle)
    results.sort(key=lambda t: t[0])
    return results


def scan_excel_files(root: Path) -> List[Path]:
    return sorted([p for p in root.rglob("*") if p.suffix.lower() in (".xlsx", ".xls")])


# -----------------------------
# Main
# -----------------------------
def main():
    file_list = scan_excel_files(ROOT_DIR)
    if not file_list:
        raise SystemExit(f"No Excel files found under {ROOT_DIR}")

    rows = []
    for f in file_list:
        cell, alpha = parse_cell_alpha(f)
        metrics = per_file_metrics(f)
        for cyc, qdis_mAh, qchg_mAh, ce in metrics:
            rows.append((cell, alpha, cyc, qdis_mAh, qchg_mAh, ce))

    out_csv = OUTDIR / "metrics_capacity_ce.csv"
    with out_csv.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["Cell", "Alpha", "Cycle", "Qdis_mAh", "Qchg_mAh", "CE_pct"])
        for r in rows:
            w.writerow(r)

    # Plot 1: all cells
    by_cell = defaultdict(list)
    for cell, alpha, cyc, qdis, qchg, ce in rows:
        by_cell[cell].append((cyc, qdis, ce))
    fig1, ax1 = plt.subplots(figsize=(9, 5.5)); ax2 = ax1.twinx()
    for cell, arr in by_cell.items():
        arr.sort(key=lambda t: t[0])
        if len(arr) > 1:
            arr = arr[:-1]
        xs, ys, ce = zip(*arr)
        ax1.plot(xs, ys, label=cell, marker="o")
        ax2.plot(xs, ce, linestyle=":", marker="*")
    ax1.set_xlabel("Cycle Number"); ax1.set_ylabel("Discharge Capacity (mAh)")
    ax2.set_ylabel("Coulombic Efficiency (%)"); ax2.set_ylim(0, 110)
    ax1.grid(False); ax1.legend(loc="best", fontsize="small", frameon=True, ncol=2)
    plt.tight_layout();plt.show(); plt.savefig(OUTDIR / "all_cells_capacity_ce.png", dpi=200, bbox_inches="tight"); plt.close(fig1)

    # Plot 2: split by alpha
    by_alpha = defaultdict(lambda: defaultdict(list))
    for cell, alpha, cyc, qdis, qchg, ce in rows:
        by_alpha[alpha][cell].append((cyc, qdis, ce))
    alphas = sorted(by_alpha.keys()); n = len(alphas); ncols = 2 if n > 1 else 1; nrows = math.ceil(n / ncols)
    fig2 = plt.figure(figsize=(7 * ncols, 4.8 * nrows))
    fig2.suptitle("Capacity & CE vs Cycle — split by cell ID (Alpha)")
    for i, a in enumerate(alphas, start=1):
        ax1 = fig2.add_subplot(nrows, ncols, i); ax2 = ax1.twinx()
        for cell, arr in sorted(by_alpha[a].items()):
            arr.sort(key=lambda t: t[0])
            if len(arr) > 1:
                arr = arr[:-1]
            xs, ys, ce = zip(*arr)
            ax1.plot(xs, ys, label=cell); ax2.plot(xs, ce, linestyle=":")
        ax1.set_title(f"Alpha {a}")
        ax1.set_xlabel("Cycle Number"); ax1.set_ylabel("Discharge Capacity (mAh)")
        ax2.set_ylabel("Coulombic Efficiency (%)"); ax2.set_ylim(0, 110)
        ax1.grid(False); ax1.legend(loc="upper center", bbox_to_anchor=(0.5, -0.18), ncol=3, fontsize="small", frameon=True)
    plt.tight_layout(rect=[0, 0, 1, 0.96]);plt.show(); plt.savefig(OUTDIR / "split_by_alpha_capacity_ce.png", dpi=200, bbox_inches="tight"); plt.close(fig2)

    print(f"Done. Wrote CSV and plots to {OUTDIR}")


if __name__ == "__main__":
    main()
