
"""
scan_charge_discharge_metrics_v2.py

Scan a ROOT_DIR recursively for Arbin-export .xlsx files.

Workflow (fast + robust):
  1) Identify "-51C" DISCHARGE files first (by filename/path patterns).
  2) Extract the cell code (e.g., IZ03) from those discharge filenames.
  3) For each discharge file, compute key discharge metrics from the time-series sheet.
  4) For that same cell code, search for CHARGE files in the scanned file index,
     then pick the best matching charge file (nearest in time to discharge start),
     and compute end-of-charge metrics.

Outputs:
  - CSV + XLSX with absolute paths and metrics.

Notes:
  - Uses openpyxl in read_only streaming mode (much faster/safer than pandas read_excel on big files).
  - Header detection handles "Date_Time", "Voltage (V)", etc.

Run:
  python scan_charge_discharge_metrics_v2.py
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any

import numpy as np
import pandas as pd
import openpyxl


# ==============================
# User config
# ==============================

# Root to scan (recursive)
ROOT_DIR = r"C:\Users\benja\Downloads"

# Where to write outputs
OUTPUT_DIR = os.path.join(ROOT_DIR, "_summary_exports")
os.makedirs(OUTPUT_DIR, exist_ok=True)

OUT_BASENAME = "charge_discharge_metrics_minus51C"

# Cutoff voltage to report "capacity at cutoff" (use None to disable)
CUTOFF_V = 2.50

# Current thresholds to detect charge/discharge segments
I_CHARGE_THRESH_A = 1e-6     # Current >  +thresh => charge
I_DISCHARGE_THRESH_A = -1e-6 # Current <  -thresh => discharge

# How many charge candidates to actually open per discharge file (ranked by file timestamp)
MAX_CHARGE_CANDIDATES_TO_OPEN = 5

# ==============================
# Filename detection helpers
# ==============================

_HYPHENS = ["\u2212", "\u2010", "\u2011", "\u2012", "\u2013", "\u2014", "\u2015"]  # minus/hyphen variants


def _norm_minus(s: str) -> str:
    """Normalize unicode minus/hyphens to ASCII '-' and lowercase."""
    for h in _HYPHENS:
        s = s.replace(h, "-")
    return s.lower()


def _tokens(s: str) -> List[str]:
    """Tokenize to alnum chunks (lowercase). 'discharge' != 'charge' here (important!)."""
    return re.findall(r"[a-z0-9]+", s.lower())


_MINUS51_RE = re.compile(r"(?<!\d)-?51c(?!\d)", flags=re.IGNORECASE)


def is_minus51_path(path: str) -> bool:
    s = _norm_minus(path)
    # common patterns: "-51C", "_-51C_", "-51C_discharges", etc.
    if _MINUS51_RE.search(s):
        return True
    # fallbacks: "minus51", "m51"
    return ("minus51" in s) or ("m51" in s)


def is_discharge_file(path: str) -> bool:
    b = os.path.basename(path)
    t = _tokens(os.path.splitext(b)[0])

    if "discharge" in t:
        return True
    if any(tok.startswith("discharge") for tok in t):
        return True
    # patterns like C10Dis -> token "c10dis"
    if any(re.fullmatch(r"c\d+dis", tok) for tok in t):
        return True
    if any("discharget" in tok for tok in t):
        return True
    return False


def is_charge_file(path: str) -> bool:
    b = os.path.basename(path)
    t = _tokens(os.path.splitext(b)[0])

    # Normal charge naming
    if "charge" in t:
        return True
    if any(tok.startswith("cccharge") or tok.startswith("charge") for tok in t):
        return True
    if any(re.fullmatch(r"c\d+chg", tok) for tok in t):
        return True

    # NEW: "FormCharge" style charge files (and common variants)
    # e.g., FormCharge, Form_Charge, FormationCharge
    if "formcharge" in t:
        return True
    if ("form" in t and "charge" in t) or ("formation" in t and "charge" in t):
        return True

    # Weak signal fallback
    if "formation" in t:
        return True

    return False



# ==============================
# Cell code extraction
# ==============================

# Prefer BL-LL-XX## pattern to avoid electrolyte codes like DT14.
_CELL_BL_RE = re.compile(r"BL[-_ ]?LL[-_ ]?([A-Z]{2}\d{2})(?!\d)", flags=re.IGNORECASE)
_CELL_GENERIC_RE = re.compile(r"(?<![A-Z0-9])([A-Z]{2})(\d{2})(?![A-Z0-9])", flags=re.IGNORECASE)


def extract_cell_code(path: str) -> Optional[str]:
    s = os.path.basename(path).upper()
    m = _CELL_BL_RE.search(s)
    if m:
        return m.group(1).upper()
    m = _CELL_GENERIC_RE.search(s)
    if m:
        return (m.group(1) + m.group(2)).upper()
    return None


def alpha_prefix(code: str) -> str:
    return code[:2]


def replicate_num(code: str) -> Optional[int]:
    try:
        return int(code[2:])
    except Exception:
        return None


# ==============================
# Workbook / header detection
# ==============================

@dataclass
class ColMap:
    dt: int
    v: int
    i: int
    q_chg: Optional[int] = None
    q_dis: Optional[int] = None
    e_chg: Optional[int] = None
    e_dis: Optional[int] = None


def _norm_label(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip().lower()
    s = s.replace("\n", " ").replace("\r", " ").replace("_", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def _find_idx(labels: List[str], must_have: List[str]) -> Optional[int]:
    """
    Return first index whose label contains all substrings in must_have.
    """
    for j, lab in enumerate(labels):
        ok = True
        for k in must_have:
            if k not in lab:
                ok = False
                break
        if ok:
            return j
    return None


def _guess_timeseries_sheet(wb: openpyxl.Workbook) -> List[str]:
    """Return sheet names in priority order for time-series data."""
    names = wb.sheetnames[:]
    def score(n: str) -> int:
        nl = n.lower()
        if "channel" in nl:
            return 0
        if "record" in nl:
            return 1
        if "data" in nl:
            return 2
        if "stat" in nl:
            return 99
        if "global" in nl:
            return 100
        return 10
    return sorted(names, key=score)


def find_timeseries_header(path: str, max_rows_scan: int = 60) -> Tuple[Optional[str], Optional[int], Optional[ColMap], Optional[str]]:
    """
    Locate the time-series sheet + header row and return (sheet, header_row, colmap, error).
    header_row is 1-indexed (Excel style).
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return None, None, None, f"openpyxl_load_failed: {e}"

    sheet_order = _guess_timeseries_sheet(wb)

    for sname in sheet_order:
        ws = wb[sname]
        try:
            for r in range(1, max_rows_scan + 1):
                row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
                labels = [_norm_label(x) for x in row]
                if not any(labels):
                    continue

                dt_idx = _find_idx(labels, ["date", "time"])
                v_idx = _find_idx(labels, ["voltage"])
                i_idx = _find_idx(labels, ["current"])

                if dt_idx is None or v_idx is None or i_idx is None:
                    continue

                q_chg = _find_idx(labels, ["charge", "capacity"])
                q_dis = _find_idx(labels, ["discharge", "capacity"])
                e_chg = _find_idx(labels, ["charge", "energy"])
                e_dis = _find_idx(labels, ["discharge", "energy"])

                wb.close()
                return sname, r, ColMap(dt=dt_idx, v=v_idx, i=i_idx, q_chg=q_chg, q_dis=q_dis, e_chg=e_chg, e_dis=e_dis), None
        except StopIteration:
            continue
        except Exception:
            continue

    try:
        wb.close()
    except Exception:
        pass
    return None, None, None, "no_timeseries_header_found"


def _to_datetime(x: Any) -> Optional[datetime]:
    if x is None:
        return None
    if isinstance(x, datetime):
        return x
    # Sometimes it's a string
    try:
        return pd.to_datetime(x).to_pydatetime()
    except Exception:
        return None


def compute_charge_end_metrics(path: str) -> Dict[str, Any]:
    """
    Compute start/end time of CHARGE segment (current > threshold),
    plus end voltage and end charge capacity/energy if available.
    """
    out = {
        "ChargePath": os.path.abspath(path),
    }
    sname, hdr, cm, err = find_timeseries_header(path)
    out["ChargeSheet"] = sname
    out["ChargeHeaderRow"] = hdr
    out["ChargeParseError"] = err
    if err or sname is None or hdr is None or cm is None:
        return out

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sname]

    chg_start = None
    chg_end = None
    v_end = None
    q_end = None
    e_end = None

    try:
        for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
            dt = _to_datetime(row[cm.dt] if cm.dt < len(row) else None)
            v = row[cm.v] if cm.v < len(row) else None
            i = row[cm.i] if cm.i < len(row) else None

            if dt is None or i is None:
                continue

            try:
                i = float(i)
            except Exception:
                continue

            if i > I_CHARGE_THRESH_A:
                if chg_start is None:
                    chg_start = dt
                chg_end = dt
                v_end = v

                if cm.q_chg is not None and cm.q_chg < len(row):
                    q_end = row[cm.q_chg]
                if cm.e_chg is not None and cm.e_chg < len(row):
                    e_end = row[cm.e_chg]
    finally:
        wb.close()

    out.update({
        "ChargeStartDateTime": chg_start,
        "ChargeEndDateTime": chg_end,
        "ChargeEndVoltage_V": float(v_end) if v_end is not None else np.nan,
        "ChargeEndCapacity_Ah": float(q_end) if q_end is not None else np.nan,
        "ChargeEndEnergy_Wh": float(e_end) if e_end is not None else np.nan,
    })
    return out


def _interp_y_at_x(x: np.ndarray, y: np.ndarray, x_target: float) -> float:
    """Linear interpolate y(x) for x_target; returns nan if not possible."""
    if x.size < 2 or not np.isfinite(x_target):
        return np.nan
    m = np.isfinite(x) & np.isfinite(y)
    x = x[m]
    y = y[m]
    if x.size < 2:
        return np.nan
    # Ensure increasing x for np.interp
    order = np.argsort(x)
    x = x[order]
    y = y[order]
    # clamp
    return float(np.interp(x_target, x, y))


def compute_minus51_discharge_metrics(path: str) -> Dict[str, Any]:
    """
    Compute key discharge stats from "-51C" discharge file:
      - start/end datetime of discharge segment (current < threshold)
      - starting voltage, ending voltage, min voltage
      - discharge capacity (Ah, mAh)
      - voltage @ 50% Q (midpoint), @10% Q, @90% Q
      - avg discharge voltage
      - discharge energy if available
      - capacity at cutoff voltage (e.g., 2.5 V) if CUTOFF_V set
    """
    out = {
        "DischargePath": os.path.abspath(path),
    }

    sname, hdr, cm, err = find_timeseries_header(path)
    out["DischargeSheet"] = sname
    out["DischargeHeaderRow"] = hdr
    out["DischargeParseError"] = err
    if err or sname is None or hdr is None or cm is None:
        return out

    if cm.q_dis is None:
        out["DischargeParseError"] = "no_discharge_capacity_column"
        return out

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sname]

    dis_start = None
    dis_end = None
    v_start = None
    v_end = None

    v_min = np.inf
    volts: List[float] = []
    caps: List[float] = []

    q_last = None
    e_last = None

    cap_at_cutoff = np.nan
    seen_cutoff = False
    prev_v = None
    prev_q = None

    try:
        for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
            dt = _to_datetime(row[cm.dt] if cm.dt < len(row) else None)
            if dt is None:
                continue

            i = row[cm.i] if cm.i < len(row) else None
            v = row[cm.v] if cm.v < len(row) else None
            q = row[cm.q_dis] if cm.q_dis < len(row) else None
            e = row[cm.e_dis] if (cm.e_dis is not None and cm.e_dis < len(row)) else None

            if i is None or v is None or q is None:
                continue
            try:
                i = float(i)
                v = float(v)
                q = float(q)
            except Exception:
                continue

            if i < I_DISCHARGE_THRESH_A:
                if dis_start is None:
                    dis_start = dt
                    v_start = v
                dis_end = dt
                v_end = v

                volts.append(v)
                caps.append(q)

                if v < v_min:
                    v_min = v

                q_last = q
                if e is not None:
                    try:
                        e_last = float(e)
                    except Exception:
                        pass

                # capacity at cutoff voltage
                if (CUTOFF_V is not None) and (not seen_cutoff):
                    if v <= CUTOFF_V:
                        # If previous point exists, interpolate capacity at exact cutoff
                        if prev_v is not None and prev_q is not None and prev_v > CUTOFF_V:
                            # interpolate q where V crosses cutoff between (prev_q, prev_v) and (q, v)
                            # Here V is decreasing; interpolate in V-space
                            try:
                                frac = (prev_v - CUTOFF_V) / (prev_v - v) if (prev_v - v) != 0 else 0.0
                                cap_at_cutoff = float(prev_q + frac * (q - prev_q))
                            except Exception:
                                cap_at_cutoff = q
                        else:
                            cap_at_cutoff = q
                        seen_cutoff = True

                prev_v = v
                prev_q = q
            else:
                # update prev for cutoff interpolation only while in discharge segment
                if dis_start is not None and (CUTOFF_V is not None) and (not seen_cutoff):
                    prev_v = v
                    prev_q = q
    finally:
        wb.close()

    if dis_start is None or q_last is None:
        return out  # no discharge segment found

    caps_arr = np.asarray(caps, dtype=float)
    volts_arr = np.asarray(volts, dtype=float)

    q_final = float(np.nanmax(caps_arr)) if caps_arr.size else float(q_last)
    q_mAh = q_final * 1000.0

    v_mid = _interp_y_at_x(caps_arr, volts_arr, 0.5 * q_final)
    v_10 = _interp_y_at_x(caps_arr, volts_arr, 0.1 * q_final)
    v_90 = _interp_y_at_x(caps_arr, volts_arr, 0.9 * q_final)

    v_avg = float(np.nanmean(volts_arr)) if volts_arr.size else np.nan

    duration_h = np.nan
    try:
        duration_h = (dis_end - dis_start).total_seconds() / 3600.0 if (dis_end and dis_start) else np.nan
    except Exception:
        pass

    out.update({
        "DischargeStartDateTime": dis_start,
        "DischargeEndDateTime": dis_end,
        "DischargeDuration_h": duration_h,
        "DischargeStartVoltage_V": float(v_start) if v_start is not None else np.nan,
        "DischargeEndVoltage_V": float(v_end) if v_end is not None else np.nan,
        "DischargeVmin_V": float(v_min) if np.isfinite(v_min) else np.nan,
        "DischargeCapacity_Ah": q_final,
        "DischargeCapacity_mAh": q_mAh,
        "VoltageAt50pctQ_V": v_mid,
        "VoltageAt10pctQ_V": v_10,
        "VoltageAt90pctQ_V": v_90,
        "AverageDischargeVoltage_V": v_avg,
        "DischargeEnergy_Wh": float(e_last) if e_last is not None else np.nan,
        "CapAtCutoffV_Ah": float(cap_at_cutoff) if np.isfinite(cap_at_cutoff) else np.nan,
        "CutoffV_V": float(CUTOFF_V) if CUTOFF_V is not None else np.nan,
    })
    return out


# ==============================
# Main scan + matching logic
# ==============================

def build_file_index(root_dir: str) -> List[str]:
    """Return list of absolute .xlsx paths under root_dir (recursive)."""
    out = []
    for dirpath, _, filenames in os.walk(root_dir):
        for fn in filenames:
            if not fn.lower().endswith(".xlsx"):
                continue
            # ignore temp Excel files
            if fn.startswith("~$"):
                continue
            out.append(os.path.abspath(os.path.join(dirpath, fn)))
    return out


def pick_best_charge_for_discharge(
    discharge_start: Optional[datetime],
    charge_paths: List[str],
) -> Optional[str]:
    """
    Fast pre-ranking based on file modification time vs discharge start.
    Returns a candidate path (does NOT open excel here).
    """
    if not charge_paths:
        return None

    # rank by closeness of file mtime to discharge_start (prefer earlier)
    scored = []
    for p in charge_paths:
        try:
            mtime = os.path.getmtime(p)
        except Exception:
            mtime = None

        if discharge_start is None or mtime is None:
            # neutral
            scored.append((0, 0, p))
            continue

        ds = discharge_start.timestamp()
        delta = ds - mtime  # positive => file modified before discharge start
        if delta >= 0:
            scored.append((0, delta, p))  # prefer smallest positive delta
        else:
            scored.append((1, abs(delta), p))  # after discharge => worse
    scored.sort(key=lambda x: (x[0], x[1]))
    return scored[0][2] if scored else None


def main():
    print(f"Scanning recursively under:\n  {ROOT_DIR}")
    all_xlsx = build_file_index(ROOT_DIR)
    print(f"Found {len(all_xlsx)} .xlsx files")

    # Index by cell code for later lookup
    by_code: Dict[str, List[str]] = {}
    discharge_files: List[str] = []

    for p in all_xlsx:
        code = extract_cell_code(p)
        if code:
            by_code.setdefault(code, []).append(p)

        # identify -51C discharge candidates (fast: name/path only)
        if is_minus51_path(p) and is_discharge_file(p):
            discharge_files.append(p)

    discharge_files = sorted(set(discharge_files))
    print(f"Found {len(discharge_files)} candidate -51C discharge files")

    # Only keep codes that actually have at least one -51C discharge file
    discharge_by_code: Dict[str, List[str]] = {}
    for p in discharge_files:
        code = extract_cell_code(p)
        if code:
            discharge_by_code.setdefault(code, []).append(p)

    codes = sorted(discharge_by_code.keys())
    print(f"Detected {len(codes)} cell codes with -51C discharges")

    rows: List[Dict[str, Any]] = []

    for code in codes:
        dis_paths = discharge_by_code[code]
        all_paths_for_code = by_code.get(code, [])

        # charge candidates: same cell code AND looks like charge AND NOT -51C discharge
        charge_candidates = [
            p for p in all_paths_for_code
            if is_charge_file(p) and (not (is_minus51_path(p) and is_discharge_file(p)))
        ]

        # De-dup and sort
        charge_candidates = sorted(set(charge_candidates))

        print(f"[{code}] -51C_discharge_files={len(dis_paths)}, charge_candidates={len(charge_candidates)}")

        for dis_p in sorted(dis_paths):
            base = os.path.basename(dis_p)
            dis_metrics = compute_minus51_discharge_metrics(dis_p)
            dis_start = dis_metrics.get("DischargeStartDateTime", None)

            # choose best charge candidate(s) to open
            best_charge_path = None
            charge_metrics: Dict[str, Any] = {
                "ChargePath": np.nan,
                "ChargeStartDateTime": pd.NaT,
                "ChargeEndDateTime": pd.NaT,
                "ChargeEndVoltage_V": np.nan,
                "ChargeEndCapacity_Ah": np.nan,
                "ChargeEndEnergy_Wh": np.nan,
                "ChargeParseError": "no_charge_candidate",
            }

            if charge_candidates:
                # rank by mtime; take top K
                ranked = []
                ds_ts = dis_start.timestamp() if isinstance(dis_start, datetime) else None
                for p in charge_candidates:
                    try:
                        mtime = os.path.getmtime(p)
                    except Exception:
                        mtime = None
                    if ds_ts is None or mtime is None:
                        rank_key = (0, 0)
                    else:
                        delta = ds_ts - mtime
                        rank_key = (0, delta) if delta >= 0 else (1, abs(delta))
                    ranked.append((rank_key, p))
                ranked.sort(key=lambda x: x[0])
                ranked_paths = [p for _, p in ranked[:MAX_CHARGE_CANDIDATES_TO_OPEN]]

                # open and pick by true charge_end_dt closeness
                best = None
                best_abs = None
                best_metrics = None
                for p in ranked_paths:
                    m = compute_charge_end_metrics(p)
                    ce = m.get("ChargeEndDateTime", None)
                    if not isinstance(ce, datetime) or not isinstance(dis_start, datetime):
                        # fallback to file mtime distance
                        try:
                            absd = abs((os.path.getmtime(p) - (ds_ts or os.path.getmtime(p))) )
                        except Exception:
                            absd = float("inf")
                    else:
                        absd = abs((dis_start - ce).total_seconds())

                    if best_abs is None or absd < best_abs:
                        best_abs = absd
                        best = p
                        best_metrics = m

                if best_metrics is not None:
                    best_charge_path = best
                    charge_metrics = best_metrics

            # Compute delta between charge end and discharge start
            delta_h = np.nan
            try:
                ce = charge_metrics.get("ChargeEndDateTime", None)
                if isinstance(ce, datetime) and isinstance(dis_start, datetime):
                    delta_h = (dis_start - ce).total_seconds() / 3600.0
            except Exception:
                pass

            row = {
                "CellCode": code,
                "Alpha": alpha_prefix(code),
                "Replicate": replicate_num(code),
                "DischargeFile": base,
                "DischargePath": os.path.abspath(dis_p),
            }
            row.update(dis_metrics)
            row.update(charge_metrics)
            row["ChargeEnd_to_DischargeStart_h"] = delta_h

            rows.append(row)

    df = pd.DataFrame(rows)

    out_csv = os.path.join(OUTPUT_DIR, f"{OUT_BASENAME}.csv")
    out_xlsx = os.path.join(OUTPUT_DIR, f"{OUT_BASENAME}.xlsx")

    if df.empty:
        print("No rows found. Check your ROOT_DIR and filename patterns.")
    else:
        # Nice ordering
        preferred = [
            "CellCode", "Alpha", "Replicate",
            "ChargePath", "ChargeStartDateTime", "ChargeEndDateTime", "ChargeEndVoltage_V", "ChargeEndCapacity_Ah",
            "DischargePath", "DischargeStartDateTime", "DischargeEndDateTime", "DischargeStartVoltage_V",
            "DischargeCapacity_Ah", "VoltageAt50pctQ_V", "CapAtCutoffV_Ah",
            "ChargeEnd_to_DischargeStart_h",
            "DischargeParseError", "ChargeParseError",
        ]
        cols = preferred + [c for c in df.columns if c not in preferred]
        df = df[cols]

        df.to_csv(out_csv, index=False)
        df.to_excel(out_xlsx, index=False)

        print(f"\nSaved CSV:\n  {out_csv}")
        print(f"Saved XLSX:\n  {out_xlsx}")
        print(f"Rows written: {len(df)}")


if __name__ == "__main__":
    main()
