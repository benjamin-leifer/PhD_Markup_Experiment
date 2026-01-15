# scan_charge_discharge_metrics.py
# Recursively scan ROOT_SCAN_DIR (all subdirs) ONCE.
# 1) Collect -51C discharge files first (by code)
# 2) Collect charge files (by code)
# 3) For each discharge file: compute -51C metrics and attach best matching charge end timestamp
#
# Uses openpyxl streaming (fast / avoids pandas read_excel slowness).

import os
import re
import math
from datetime import datetime, timedelta
from collections import defaultdict
from typing import Optional, Dict, Any, List, Tuple

import pandas as pd
from openpyxl import load_workbook


# =========================
# CONFIG (edit these)
# =========================
# This WILL scan root + all subdirectories recursively.
ROOT_SCAN_DIR = r"C:\Users\benja\Downloads\Dilute THF Data\11_25_25"

# Output
OUT_DIR = os.path.join(ROOT_SCAN_DIR, "_summary_exports")
OUT_BASENAME = "charge_discharge_metrics_minus51C"

# -51C tag heuristics
TEMP_TAGS = ["-51c", "-51", "minus51", "m51"]

# File name heuristics
DISCHARGE_HINTS = ["discharge", "c10dis", "discharget", "_dis_", "_discharge", "dis_"]
CHARGE_HINTS = ["charge", "c10chg", "_chg", "_charge", "chg_"]

# Current threshold (A) to classify charge/discharge rows
I_CHG_MIN = 1e-6
I_DIS_MAX = -1e-6

# Header scan limits
MAX_HEADER_ROWS_TO_SCAN = 80
MAX_SHEETS_TO_SCAN = 8  # scan first N sheets for a usable header


# =========================
# Cell code extraction
# =========================
# Prefer strict BL-LL-XX## match to avoid false positives like "DT14"
CELL_CODE_RE_STRICT = re.compile(r"BL-LL-([A-Z]{2}\d{2})", re.IGNORECASE)
# Fallback: match XX## when bounded by non-alphanumerics
CELL_CODE_RE_GENERIC = re.compile(r"(?<![A-Z0-9])([A-Z]{2}\d{2})(?!\d)", re.IGNORECASE)


def extract_cell_code(path: str) -> Optional[str]:
    m = CELL_CODE_RE_STRICT.search(path)
    if m:
        return m.group(1).upper()

    base = os.path.basename(path)
    hits = CELL_CODE_RE_GENERIC.findall(base)
    if hits:
        return hits[-1].upper()

    hits = CELL_CODE_RE_GENERIC.findall(path)
    if hits:
        return hits[-1].upper()

    return None


# =========================
# File filters
# =========================
def _has_any(s: str, keys: List[str]) -> bool:
    s = s.lower()
    return any(k in s for k in keys)


def looks_like_minus51_discharge(path: str) -> bool:
    s = path.lower()
    b = os.path.basename(path).lower()
    is_m51 = _has_any(s, TEMP_TAGS)
    is_dis = _has_any(b, DISCHARGE_HINTS)
    # IMPORTANT: "discharge" contains "charge"
    not_charge = ("charge" not in b)
    return is_m51 and is_dis and not_charge


def looks_like_charge(path: str) -> bool:
    b = os.path.basename(path).lower()
    if "discharge" in b:
        return False
    return _has_any(b, CHARGE_HINTS)


# =========================
# Excel streaming helpers
# =========================
def _to_float(x) -> float:
    try:
        if x is None:
            return float("nan")
        if isinstance(x, bool):
            return float("nan")
        return float(x)
    except Exception:
        return float("nan")


def _parse_datetime(val) -> Optional[datetime]:
    if val is None:
        return None

    if isinstance(val, datetime):
        return val

    # Excel serial date (days since 1899-12-30)
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        fv = float(val)
        if 20000.0 < fv < 70000.0:  # rough bounds for modern dates
            base = datetime(1899, 12, 30)
            return base + timedelta(days=fv)

    # String parse
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        fmts = [
            "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
            "%Y/%m/%d %H:%M:%S",
        ]
        for f in fmts:
            try:
                return datetime.strptime(s, f)
            except Exception:
                pass
        try:
            ts = pd.to_datetime(s, errors="coerce")
            if pd.isna(ts):
                return None
            return ts.to_pydatetime()
        except Exception:
            return None

    return None


def _normalize_header_cell(x) -> str:
    if x is None:
        return ""
    return str(x).strip().lower().replace("\n", " ").replace("\r", " ")


def _find_best_sheet_and_header(path: str) -> Tuple[Optional[str], Optional[int], Optional[List[str]]]:
    """
    Return (sheet_name, header_row_idx_1based, header_labels_lower)
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None, None, None

    try:
        sheetnames = wb.sheetnames[:MAX_SHEETS_TO_SCAN]
        for sh in sheetnames:
            ws = wb[sh]
            for r_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=MAX_HEADER_ROWS_TO_SCAN, values_only=True),
                start=1
            ):
                labels = [_normalize_header_cell(c) for c in row]
                has_v = any("volt" in c or c in ("ewe/v",) for c in labels)
                has_i = any("current" in c or c in ("i/a",) for c in labels)
                has_dt = any(("date" in c and "time" in c) or c in ("date time", "datetime") for c in labels)
                has_cap = any("capacity" in c for c in labels)

                if has_v and has_i and (has_dt or has_cap):
                    return sh, r_idx, labels
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return None, None, None


def _header_index_map(labels_lower: List[str]) -> Dict[str, int]:
    idx = {}

    def pick(preds: List[str], contains_any: List[str] = None) -> Optional[int]:
        for j, lab in enumerate(labels_lower):
            if any(lab == p for p in preds):
                return j
        if contains_any:
            for j, lab in enumerate(labels_lower):
                if any(k in lab for k in contains_any):
                    return j
        return None

    j_dt = pick(["date time", "datetime", "date/time"], contains_any=["date time", "datetime"])
    if j_dt is None:
        j_dt = pick([], contains_any=["date", "time"])
    if j_dt is not None:
        idx["dt"] = j_dt

    j_i = pick(["current (a)", "current(a)", "i (a)", "i(a)"], contains_any=["current"])
    if j_i is not None:
        idx["i"] = j_i

    j_v = pick(["voltage (v)", "voltage(v)", "ewe/v", "v (v)", "v(v)"], contains_any=["volt", "ewe/v"])
    if j_v is not None:
        idx["v"] = j_v

    j_qdis = pick([], contains_any=["discharge capacity"])
    if j_qdis is not None:
        idx["q_dis"] = j_qdis

    j_qchg = pick([], contains_any=["charge capacity"])
    if j_qchg is not None:
        idx["q_chg"] = j_qchg

    j_q = pick(["capacity (ah)", "capacity(ah)", "capacity (mah)", "capacity(mah)"], contains_any=["capacity"])
    if j_q is not None:
        idx["q"] = j_q

    return idx


def _stream_rows(path: str, sheet: str, header_row_1based: int):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    try:
        for row in ws.iter_rows(min_row=header_row_1based + 1, values_only=True):
            yield row
    finally:
        try:
            wb.close()
        except Exception:
            pass


# =========================
# Metric computations
# =========================
def compute_charge_end_metrics(path: str) -> Dict[str, Any]:
    sh, hdr, labels = _find_best_sheet_and_header(path)
    if sh is None:
        return {"ChargeParseOK": False, "ChargeParseError": "No usable sheet/header found"}

    idx = _header_index_map(labels)
    if "i" not in idx or "v" not in idx:
        return {"ChargeParseOK": False, "ChargeParseError": "Missing Current/Voltage columns"}

    dt_last = None
    v_last = float("nan")
    q_last = float("nan")

    for row in _stream_rows(path, sh, hdr):
        try:
            i = _to_float(row[idx["i"]]) if idx["i"] < len(row) else float("nan")
            if not math.isfinite(i) or i <= I_CHG_MIN:
                continue

            dt = _parse_datetime(row[idx["dt"]]) if "dt" in idx and idx["dt"] < len(row) else None
            v = _to_float(row[idx["v"]]) if idx["v"] < len(row) else float("nan")

            q = float("nan")
            if "q_chg" in idx and idx["q_chg"] < len(row):
                q = _to_float(row[idx["q_chg"]])
            elif "q" in idx and idx["q"] < len(row):
                q = _to_float(row[idx["q"]])

            if dt is not None:
                dt_last = dt
            if math.isfinite(v):
                v_last = v
            if math.isfinite(q):
                q_last = q

        except Exception:
            continue

    return {
        "ChargeParseOK": True,
        "ChargeEndDateTime": dt_last,
        "ChargeEndVoltage_V": v_last if math.isfinite(v_last) else None,
        "ChargeEndCapacity_Ah": q_last if math.isfinite(q_last) else None,
    }


def compute_discharge_minus51_metrics(path: str) -> Dict[str, Any]:
    sh, hdr, labels = _find_best_sheet_and_header(path)
    if sh is None:
        return {"DischargeParseOK": False, "DischargeParseError": "No usable sheet/header found"}

    idx = _header_index_map(labels)
    if "i" not in idx or "v" not in idx:
        return {"DischargeParseOK": False, "DischargeParseError": "Missing Current/Voltage columns"}

    dt0 = None
    dt1 = None
    v0 = float("nan")
    v1 = float("nan")

    q0 = None
    q_last = None

    q_list: List[float] = []
    v_list: List[float] = []

    for row in _stream_rows(path, sh, hdr):
        try:
            i = _to_float(row[idx["i"]]) if idx["i"] < len(row) else float("nan")
            if not math.isfinite(i) or i >= I_DIS_MAX:
                continue

            dt = _parse_datetime(row[idx["dt"]]) if "dt" in idx and idx["dt"] < len(row) else None
            v = _to_float(row[idx["v"]]) if idx["v"] < len(row) else float("nan")

            q = float("nan")
            if "q_dis" in idx and idx["q_dis"] < len(row):
                q = _to_float(row[idx["q_dis"]])
            elif "q" in idx and idx["q"] < len(row):
                q = _to_float(row[idx["q"]])

            if dt0 is None and dt is not None:
                dt0 = dt
            if dt is not None:
                dt1 = dt

            if not math.isfinite(v0) and math.isfinite(v):
                v0 = v
            if math.isfinite(v):
                v1 = v

            if math.isfinite(q):
                if q0 is None:
                    q0 = q
                q_last = q
                q_list.append(q)
                v_list.append(v if math.isfinite(v) else float("nan"))

        except Exception:
            continue

    if q_last is None or q0 is None or len(q_list) < 2:
        return {
            "DischargeParseOK": False,
            "DischargeParseError": "No discharge rows with capacity detected (Current<0 and capacity col present)",
            "DischargeStartDateTime": dt0,
            "DischargeEndDateTime": dt1,
        }

    qn = [q - q0 for q in q_list]
    q_total = qn[-1]

    v_mid = None
    if q_total > 0 and len(qn) > 1:
        target = 0.5 * q_total
        k = min(range(len(qn)), key=lambda i: abs(qn[i] - target))
        v_mid = v_list[k] if math.isfinite(v_list[k]) else None

    energy_Wh = None
    v_avg = None
    if q_total > 0 and len(qn) > 2:
        e = 0.0
        q_acc = 0.0
        for i in range(1, len(qn)):
            q1 = qn[i - 1]
            q2 = qn[i]
            dq = q2 - q1
            if not (math.isfinite(q1) and math.isfinite(q2) and dq > 0):
                continue
            v_a = v_list[i - 1]
            v_b = v_list[i]
            if not (math.isfinite(v_a) and math.isfinite(v_b)):
                continue
            e += 0.5 * (v_a + v_b) * dq
            q_acc += dq
        if q_acc > 0:
            energy_Wh = e
            v_avg = e / q_acc

    return {
        "DischargeParseOK": True,
        "DischargeStartDateTime": dt0,
        "DischargeEndDateTime": dt1,
        "DischargeStartV_V": v0 if math.isfinite(v0) else None,
        "DischargeEndV_V": v1 if math.isfinite(v1) else None,
        "DischargeCapacity_Ah": float(q_total),
        "DischargeCapacity_mAh": float(q_total) * 1000.0,
        "Voltage_at_50pctQ_V": v_mid,
        "AvgDischargeV_V": v_avg,
        "Energy_Wh": energy_Wh,
    }


# =========================
# Charge selection
# =========================
def pick_best_charge_file(charge_candidates: List[str], discharge_path: str) -> Optional[str]:
    if not charge_candidates:
        return None

    try:
        t_dis = os.path.getmtime(discharge_path)
    except Exception:
        t_dis = None

    if t_dis is not None:
        eligible = []
        for p in charge_candidates:
            try:
                t = os.path.getmtime(p)
                if t <= t_dis:
                    eligible.append((t, p))
            except Exception:
                pass
        if eligible:
            eligible.sort()
            return eligible[-1][1]

        best = None
        best_dt = None
        for p in charge_candidates:
            try:
                t = os.path.getmtime(p)
                d = abs(t - t_dis)
                if best_dt is None or d < best_dt:
                    best_dt = d
                    best = p
            except Exception:
                pass
        return best

    return max(charge_candidates, key=lambda p: os.path.getmtime(p))


# =========================
# MAIN
# =========================
def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    # Single recursive walk of ROOT_SCAN_DIR
    by_cell_dis: Dict[str, List[str]] = defaultdict(list)
    by_cell_chg_all: Dict[str, List[str]] = defaultdict(list)

    total_xlsx = 0
    for r, _dirs, files in os.walk(ROOT_SCAN_DIR):
        for fn in files:
            if not fn.lower().endswith(".xlsx") or fn.startswith("~$"):
                continue
            total_xlsx += 1
            path = os.path.join(r, fn)
            code = extract_cell_code(path)
            if not code:
                continue

            if looks_like_minus51_discharge(path):
                by_cell_dis[code].append(path)
            elif looks_like_charge(path):
                by_cell_chg_all[code].append(path)

    discharge_codes = sorted(by_cell_dis.keys())

    print(f"Scanned recursively under:\n  {ROOT_SCAN_DIR}")
    print(f"Found {total_xlsx} .xlsx files total")
    print(f"Detected {len(discharge_codes)} cell codes with -51C discharges.")
    print()

    if not discharge_codes:
        print("No -51C discharge files matched. Check TEMP_TAGS / DISCHARGE_HINTS or directory.")
        out_csv = os.path.join(OUT_DIR, OUT_BASENAME + ".csv")
        pd.DataFrame([]).to_csv(out_csv, index=False)
        print(f"Saved empty CSV:\n  {out_csv}")
        return

    rows: List[Dict[str, Any]] = []
    charge_cache: Dict[str, Dict[str, Any]] = {}
    dis_cache: Dict[str, Dict[str, Any]] = {}

    for code in discharge_codes:
        dis_list = sorted(by_cell_dis[code])
        chg_list = sorted(by_cell_chg_all.get(code, []))

        print(f"[{code}] -51C_discharge_files={len(dis_list)}, charge_files={len(chg_list)}")

        for dis_path in dis_list:
            if dis_path in dis_cache:
                dm = dis_cache[dis_path]
            else:
                try:
                    dm = compute_discharge_minus51_metrics(dis_path)
                except Exception as e:
                    dm = {"DischargeParseOK": False, "DischargeParseError": str(e)}
                dis_cache[dis_path] = dm

            chg_path = pick_best_charge_file(chg_list, dis_path)

            cm: Dict[str, Any] = {}
            if chg_path:
                if chg_path in charge_cache:
                    cm = charge_cache[chg_path]
                else:
                    try:
                        cm = compute_charge_end_metrics(chg_path)
                    except Exception as e:
                        cm = {"ChargeParseOK": False, "ChargeParseError": str(e)}
                    charge_cache[chg_path] = cm

            row = {
                "CellCode": code,
                "DischargePath": os.path.abspath(dis_path),
                "ChargePath": os.path.abspath(chg_path) if chg_path else None,
                "DischargeFile_mtime": os.path.getmtime(dis_path),
                "ChargeFile_mtime": os.path.getmtime(chg_path) if chg_path else None,
            }
            row.update(cm)
            row.update(dm)

            try:
                row["DischargeDuration_hr"] = (
                    (row["DischargeEndDateTime"] - row["DischargeStartDateTime"]).total_seconds() / 3600.0
                    if row.get("DischargeStartDateTime") and row.get("DischargeEndDateTime") else None
                )
            except Exception:
                row["DischargeDuration_hr"] = None

            rows.append(row)

    df = pd.DataFrame(rows)

    for c in ["ChargeEndDateTime", "DischargeStartDateTime", "DischargeEndDateTime"]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")

    out_csv = os.path.join(OUT_DIR, OUT_BASENAME + ".csv")
    out_xlsx = os.path.join(OUT_DIR, OUT_BASENAME + ".xlsx")
    df.to_csv(out_csv, index=False)
    df.to_excel(out_xlsx, index=False)

    print(f"\nSaved CSV:\n  {out_csv}\nSaved XLSX:\n  {out_xlsx}\nRows written: {len(df)}")


if __name__ == "__main__":
    main()
