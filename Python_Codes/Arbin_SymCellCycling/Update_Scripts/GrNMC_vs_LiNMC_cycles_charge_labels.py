import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from pathlib import Path
from matplotlib.lines import Line2D

# -------------------------
# SETTINGS
# -------------------------
ACTIVE_MASS_G = 0.01597586  # g

DQ01_PATH = "/mnt/data/BL-LL-DQ01_RT_Rate_Test_Channel_16_Wb_1.xlsx"  # Gr|NMC
DP01_PATH = "/mnt/data/BL-LL-DP01_RT_Rate_Test_Channel_11_Wb_1.xlsx"  # Li|NMC

CYCLES_TO_PLOT = [1, 2, 5, 20, 50]
DQ_VMIN = 2.5  # clip Gr|NMC curves

OUT_PATH = "/mnt/data/GrNMC_vs_LiNMC_charge_labeled_cycles.png"

# Linewidth encodes cycle number
LW_BY_CYCLE = {1: 1.6, 2: 2.0, 5: 2.6, 20: 3.2, 50: 4.0}

# -------------------------
# HELPERS
# -------------------------
def extract_up_to_cycle_openpyxl(fp: str, max_cycle: int) -> pd.DataFrame:
    wb = load_workbook(fp, read_only=True, data_only=True)
    ws = wb.worksheets[1]  # second sheet

    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    def idx(col):
        return header.index(col)

    i_cycle = idx("Cycle Index")
    i_V = idx("Voltage (V)")
    i_I = idx("Current (A)")
    i_Qchg = idx("Charge Capacity (Ah)")
    i_Qdis = idx("Discharge Capacity (Ah)")

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[i_cycle] is None:
            continue
        cyc = int(r[i_cycle])
        if cyc > max_cycle:
            break

        V = r[i_V]
        I = r[i_I]
        Qc = r[i_Qchg] or 0.0
        Qd = r[i_Qdis] or 0.0
        if V is None or I is None:
            continue

        rows.append((cyc, float(V), float(I), float(Qc), float(Qd)))

    wb.close()
    return pd.DataFrame(rows, columns=["cycle", "V", "I", "Qchg_Ah", "Qdis_Ah"])


def prep_halfcycle(df, cycle, direction, mass_g, vmin=None):
    d = df[df["cycle"] == cycle].copy()

    if direction == "charge":
        h = d[d["I"] > 0].copy()
        cap = h["Qchg_Ah"]
    else:
        h = d[d["I"] < 0].copy()
        cap = h["Qdis_Ah"]

    if len(h) == 0:
        return h

    h = h.assign(cap_Ah=cap).sort_values("cap_Ah")

    if vmin is not None:
        h = h[h["V"] >= vmin].copy()
        if len(h) == 0:
            return h

    h["cap_mAh_g"] = ((h["cap_Ah"] - h["cap_Ah"].iloc[0]) * 1000) / mass_g
    return h[["cap_mAh_g", "V"]]


def plot_cell(ax, df, color, title, vmin=None):
    for cyc in CYCLES_TO_PLOT:
        lw = LW_BY_CYCLE[cyc]

        # Charge (solid)
        chg = prep_halfcycle(df, cyc, "charge", ACTIVE_MASS_G, vmin=vmin)
        if len(chg):
            ax.plot(
                chg["cap_mAh_g"],
                chg["V"],
                color=color,
                linewidth=lw,
                linestyle="-",
            )

            # label ONLY at top of charge
            ax.text(
                chg["cap_mAh_g"].iloc[-1],
                chg["V"].iloc[-1] + 0.03,
                "50 (C/2)" if cyc == 50 else f"{cyc}",
                fontsize=13,
                fontweight="bold",
                color=color,
                ha="left",
                va="bottom",
            )

        # Discharge (dashed)
        dis = prep_halfcycle(df, cyc, "discharge", ACTIVE_MASS_G, vmin=vmin)
        if len(dis):
            ax.plot(
                dis["cap_mAh_g"],
                dis["V"],
                color=color,
                linewidth=lw,
                linestyle="--",
            )

    ax.set_title(title, fontsize=18, fontweight="bold")
    ax.set_xlabel("Discharge Capacity (mAh/g)", fontsize=15)
    ax.set_ylabel("Voltage (V)", fontsize=15)
    ax.tick_params(axis="both", labelsize=13)
    ax.set_xlim(left=0)


# -------------------------
# LOAD DATA
# -------------------------
max_cycle = max(CYCLES_TO_PLOT)
gr_nmc = extract_up_to_cycle_openpyxl(DQ01_PATH, max_cycle)
li_nmc = extract_up_to_cycle_openpyxl(DP01_PATH, max_cycle)

# -------------------------
# PLOT
# -------------------------
fig, axes = plt.subplots(1, 2, figsize=(14, 6), sharey=True)

plot_cell(
    axes[0],
    gr_nmc,
    color="black",
    title="Gr|NMC",
    vmin=DQ_VMIN,
)

plot_cell(
    axes[1],
    li_nmc,
    color="green",
    title="Li|NMC",
    vmin=None,
)

# Voltage headroom
for ax in axes:
    ymin, ymax = ax.get_ylim()
    ax.set_ylim(ymin, ymax + 0.15)
    ax.tick_params(direction="in", top=True, right=True)

# -------------------------
# LEGEND
# -------------------------
cycle_handles = [
    Line2D([0], [0], color="k", lw=LW_BY_CYCLE[c], label=f"Cycle {c}" if c != 50 else "Cycle 50 (C/2)")
    for c in CYCLES_TO_PLOT
]
dir_handles = [
    Line2D([0], [0], color="k", lw=2.5, linestyle="-", label="Charge"),
    Line2D([0], [0], color="k", lw=2.5, linestyle="--", label="Discharge"),
]

fig.legend(
    cycle_handles + dir_handles,
    [h.get_label() for h in cycle_handles + dir_handles],
    loc="lower center",
    ncol=4,
    frameon=False,
    fontsize=12,
)

plt.tight_layout(rect=[0, 0.08, 1, 1])
fig.savefig(OUT_PATH, dpi=300)
plt.show()

print(f"Saved to: {OUT_PATH}")