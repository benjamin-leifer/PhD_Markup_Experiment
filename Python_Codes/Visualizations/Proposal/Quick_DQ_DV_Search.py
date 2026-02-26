"""
Scan a directory for Excel files whose filename/path contains:
  1) a known CELL PREFIX (e.g., "AA") from your lookup table, AND
  2) one of: "Form", "Rate", "HiFi" (case-insensitive)

Then:
  - Extract explicit cell_code like AA01 (prefix + digits)
  - Filter out prefixes "before HC" using your lookup's Number column
  - Drop duplicates (same cell_code + tags), keep newest modified_time

Export an Excel summary with:
  cell_prefix | cell_code | replicate | electrolyte | tags | file_path | filename | modified_time
"""

from __future__ import annotations

import re
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# -------------------------
# SETTINGS (EDIT THESE)
# -------------------------
SEARCH_ROOT = Path(r"C:\Users\benja\Downloads\Dilute THF Data\11_25_25\DPE Room Temperature Data")  # folder to scan recursively
LOOKUP_XLSX = Path(r"C:\Users\benja\OneDrive - Northeastern University\Spring 2025 Cell List.xlsx")
LOOKUP_SHEET = 0  # 0 = first sheet

# Your lookup headers:
LOOKUP_CELL_COL = "Cell Code"
LOOKUP_ELECTROLYTE_COL = "Electrolyte"
LOOKUP_NUMBER_COL = "Number"  # used for "drop anything before HC (numerically)"

# Threshold
MIN_PREFIX = "HC"

OUTPUT_XLSX = Path(r"C:\Users\benja\Downloads\cell_file_index.xlsx")

TAG_KEYWORDS = ("form", "rate", "hifi")
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}


# -------------------------
# HELPERS
# -------------------------
def code_to_base26(code: str) -> int:
    """Fallback ordering if lookup 'Number' isn't usable."""
    code = re.sub(r"[^A-Z]", "", str(code).upper())
    if not code:
        return -1
    v = 0
    for ch in code:
        v = v * 26 + (ord(ch) - ord("A"))
    return v


def load_lookup_filtered(lookup_path: Path) -> tuple[dict[str, str], set[str]]:
    """
    Returns:
      electrolyte_by_prefix: dict[prefix -> electrolyte]
      allowed_prefixes: prefixes with Number >= Number(MIN_PREFIX) (or base-26 fallback)
    """
    df = pd.read_excel(lookup_path, sheet_name=LOOKUP_SHEET)

    required = {LOOKUP_CELL_COL, LOOKUP_ELECTROLYTE_COL}
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Lookup table missing required columns: {missing}. Found: {list(df.columns)}")

    use_number = LOOKUP_NUMBER_COL in df.columns

    cols = [LOOKUP_CELL_COL, LOOKUP_ELECTROLYTE_COL] + ([LOOKUP_NUMBER_COL] if use_number else [])
    df = df[cols].copy()

    df[LOOKUP_CELL_COL] = df[LOOKUP_CELL_COL].astype(str).str.strip().str.upper()
    df[LOOKUP_ELECTROLYTE_COL] = df[LOOKUP_ELECTROLYTE_COL].astype(str).str.strip()

    # drop blanks
    df = df[df[LOOKUP_CELL_COL].str.len() > 0]

    # Determine threshold
    allowed_prefixes: set[str]
    if use_number:
        df[LOOKUP_NUMBER_COL] = pd.to_numeric(df[LOOKUP_NUMBER_COL], errors="coerce")
        hc_rows = df[df[LOOKUP_CELL_COL] == MIN_PREFIX]
        hc_num = None
        if not hc_rows.empty:
            v = hc_rows[LOOKUP_NUMBER_COL].iloc[0]
            if pd.notna(v):
                hc_num = float(v)

        if hc_num is not None:
            df = df[pd.notna(df[LOOKUP_NUMBER_COL])]
            df = df[df[LOOKUP_NUMBER_COL] >= hc_num]
            allowed_prefixes = set(df[LOOKUP_CELL_COL].unique())
        else:
            # Fallback: base-26 compare if HC not found / Number unusable
            thr = code_to_base26(MIN_PREFIX)
            allowed_prefixes = {c for c in df[LOOKUP_CELL_COL].unique() if code_to_base26(c) >= thr}
            df = df[df[LOOKUP_CELL_COL].isin(allowed_prefixes)]
    else:
        thr = code_to_base26(MIN_PREFIX)
        allowed_prefixes = {c for c in df[LOOKUP_CELL_COL].unique() if code_to_base26(c) >= thr}
        df = df[df[LOOKUP_CELL_COL].isin(allowed_prefixes)]

    # Map electrolyte (first occurrence wins)
    df = df.drop_duplicates(subset=[LOOKUP_CELL_COL], keep="first")
    electrolyte_by_prefix = df.set_index(LOOKUP_CELL_COL)[LOOKUP_ELECTROLYTE_COL].to_dict()

    return electrolyte_by_prefix, allowed_prefixes


def extract_tags(text: str) -> list[str]:
    t = text.lower()
    hits = [k for k in TAG_KEYWORDS if k in t]
    pretty = {"form": "Form", "rate": "Rate", "hifi": "HiFi"}
    return [pretty[h] for h in hits]


def build_prefix_pattern(prefixes: set[str]) -> re.Pattern:
    """
    Match any allowed prefix ONLY when followed by optional separators and then a digit.
    e.g. AA01, BL-LL-AA01, AA_01, AA-01
    """
    esc = sorted((re.escape(p) for p in prefixes if p), key=len, reverse=True)
    if not esc:
        return re.compile(r"(?!x)x")
    pattern = (
        r"(?<![A-Z0-9])"
        r"(" + "|".join(esc) + r")"
        r"(?=(?:[_\- ]*\d))"
    )
    return re.compile(pattern, flags=re.IGNORECASE)


def extract_full_cell_code(path_text: str, prefix: str) -> tuple[str | None, int | None]:
    """
    Extract explicit code like AA01 from the path, given prefix 'AA'.
    Returns (cell_code, replicate_int).
    """
    # prefix + optional separators + digits
    pat = re.compile(rf"(?<![A-Z0-9])({re.escape(prefix)})(?:[_\- ]*)?(\d{{1,3}})", re.IGNORECASE)
    m = pat.search(path_text.upper())
    if not m:
        return None, None
    digits = m.group(2)
    rep = int(digits)
    # Zero-pad 1-digit to 2 (AA1 -> AA01), keep 2+ digits as-is
    digits_norm = digits.zfill(2) if len(digits) == 1 else digits
    return f"{prefix}{digits_norm}", rep


def autosize_and_format_excel(path: Path, freeze_panes: str = "A2") -> None:
    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = freeze_panes
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col_cells:
            v = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 120)

    wb.save(path)


# -------------------------
# MAIN
# -------------------------
def main() -> None:
    electrolyte_by_prefix, allowed_prefixes = load_lookup_filtered(LOOKUP_XLSX)
    prefix_pat = build_prefix_pattern(allowed_prefixes)

    rows = []
    n_excel = 0
    n_tagged = 0
    n_matched_prefix = 0
    n_with_full_code = 0

    for p in SEARCH_ROOT.rglob("*"):
        if not p.is_file():
            continue
        if p.suffix.lower() not in EXCEL_SUFFIXES:
            continue

        n_excel += 1
        path_str = str(p)

        tags = extract_tags(path_str)
        if not tags:
            continue
        n_tagged += 1

        m = prefix_pat.search(path_str.upper())
        if not m:
            continue
        cell_prefix = m.group(1).upper()
        n_matched_prefix += 1

        cell_code, rep = extract_full_cell_code(path_str, cell_prefix)
        if cell_code is None:
            # If you want to keep prefix-only hits, change this to not-continue
            continue
        n_with_full_code += 1

        rows.append(
            {
                "cell_prefix": cell_prefix,
                "cell_code": cell_code,
                "replicate": rep,
                "electrolyte": electrolyte_by_prefix.get(cell_prefix, ""),
                "tags": ", ".join(tags),
                "file_path": path_str,
                "filename": p.name,
                "modified_time": datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds"),
            }
        )

    out_df = pd.DataFrame(rows)

    if not out_df.empty:
        # Keep newest file for duplicates of (cell_code, tags)
        out_df["modified_time_dt"] = pd.to_datetime(out_df["modified_time"], errors="coerce")
        out_df = out_df.sort_values(["modified_time_dt", "filename"], ascending=[False, True], kind="stable")

        # Drop exact duplicate paths first (cheap safety)
        out_df = out_df.drop_duplicates(subset=["file_path"], keep="first")

        # Drop duplicates by (cell_code, tags), keep newest
        out_df = out_df.drop_duplicates(subset=["cell_code", "tags"], keep="first")

        # Final sort for readability
        out_df = out_df.sort_values(["cell_prefix", "replicate", "tags", "filename"], kind="stable")
        out_df = out_df.drop(columns=["modified_time_dt"], errors="ignore")

    out_df.to_excel(OUTPUT_XLSX, index=False)
    autosize_and_format_excel(OUTPUT_XLSX)

    print(f"Scanned Excel files: {n_excel}")
    print(f"With Form/Rate/HiFi tag: {n_tagged}")
    print(f"Matched allowed prefix (>= {MIN_PREFIX}): {n_matched_prefix}")
    print(f"With explicit cell_code (e.g., AA01): {n_with_full_code}")
    print(f"Rows written (after dedupe): {len(out_df)}")
    print(f"Saved: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
